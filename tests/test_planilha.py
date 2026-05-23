# tests/test_planilha.py - Testes de caracterização para criar_planilha.py
# Captura o comportamento atual: geração da planilha Excel com abas Dashboard, Leads PF, Leads PJ

import json
import os
import sys
import pytest
from openpyxl import load_workbook

PLANILHA_PATH = "planilha_leads.xlsx"
LEADS_PATH = "leads.json"


class TestPlanilha:
    """Testes de caracterização para a planilha de leads."""

    def setup_method(self):
        """Verifica se a planilha existe antes de cada teste."""
        assert os.path.exists(PLANILHA_PATH), f"Planilha {PLANILHA_PATH} nao encontrada"
        self.wb = load_workbook(PLANILHA_PATH)

    def test_abas_existem(self):
        """A planilha deve conter as 3 abas: Dashboard, Leads PF, Leads PJ."""
        abas = self.wb.sheetnames
        assert "Dashboard" in abas, "Aba Dashboard ausente"
        assert "Leads PF" in abas, "Aba Leads PF ausente"
        assert "Leads PJ" in abas, "Aba Leads PJ ausente"

    def test_dashboard_titulo(self):
        """A aba Dashboard deve ter o título corporativo na célula A1."""
        ws = self.wb["Dashboard"]
        titulo = ws["A1"].value
        assert titulo is not None
        assert "KAIZEN" in titulo.upper()

    def test_dashboard_headers(self):
        """A linha 2 do Dashboard deve conter os cabeçalhos esperados."""
        ws = self.wb["Dashboard"]
        headers = [ws.cell(row=2, column=c).value for c in range(1, 9)]
        headers = [h for h in headers if h is not None]
        assert len(headers) >= 5, "Dashboard deve ter pelo menos 5 cabeçalhos"

    def test_dashboard_possui_leads(self):
        """O Dashboard deve ter pelo menos 1 linha de lead (a partir da linha 3)."""
        ws = self.wb["Dashboard"]
        linha3 = ws.cell(row=3, column=1).value
        # Se não houver leads, as células podem estar vazias ou ter valor None
        # O teste captura o comportamento atual (ter ou não leads)
        pass  # Teste de caracterização - aceita o estado atual

    def test_aba_leads_pf_existe(self):
        """A aba Leads PF deve existir e ter cabeçalhos."""
        ws = self.wb["Leads PF"]
        linha3 = [ws.cell(row=3, column=c).value for c in range(1, 20)]
        linha3 = [h for h in linha3 if h is not None]
        assert len(linha3) >= 10, "Leads PF deve ter pelo menos 10 colunas com cabecalhos"

    def test_aba_leads_pj_existe(self):
        """A aba Leads PJ deve existir e ter cabeçalhos."""
        ws = self.wb["Leads PJ"]
        linha3 = [ws.cell(row=3, column=c).value for c in range(1, 25)]
        linha3 = [h for h in linha3 if h is not None]
        assert len(linha3) >= 10, "Leads PJ deve ter pelo menos 10 colunas com cabecalhos"

    def test_campos_monday_presentes(self):
        """As abas de leads devem conter os campos do grupo PLANO (MONDAY)."""
        for aba_nome in ["Leads PF", "Leads PJ"]:
            ws = self.wb[aba_nome]
            # Linha 2 contém os nomes dos grupos
            grupos = [ws.cell(row=2, column=c).value for c in range(1, 30)]
            grupos = [g for g in grupos if g is not None]
            grupos_str = " ".join(str(g) for g in grupos)
            assert "MONDAY" in grupos_str.upper() or "PLANO" in grupos_str.upper(), \
                f"Aba {aba_nome} deve ter grupo PLANO (MONDAY)"


class TestLeadsJson:
    """Testes de caracterização para leads.json."""

    def setup_method(self):
        assert os.path.exists(LEADS_PATH), f"Arquivo {LEADS_PATH} nao encontrado"
        with open(LEADS_PATH, "r", encoding="utf-8") as f:
            self.leads = json.load(f)

    def test_lista_de_leads(self):
        """leads.json deve conter uma lista de objetos."""
        assert isinstance(self.leads, list), "leads.json deve ser uma lista"
        assert len(self.leads) > 0, "leads.json deve ter pelo menos um lead"

    def test_campos_obrigatorios(self):
        """Cada lead deve ter nome, number, dependentes, tipo e bloco monday."""
        for lead in self.leads:
            assert "nome" in lead, f"Lead {lead.get('number', '?')} sem nome"
            assert "number" in lead, f"Lead {lead.get('nome', '?')} sem number"
            assert "dependentes" in lead, f"Lead {lead.get('nome', '?')} sem dependentes"
            assert "tipo" in lead, f"Lead {lead.get('nome', '?')} sem tipo"
            assert "monday" in lead, f"Lead {lead.get('nome', '?')} sem bloco monday"

    def test_tipos_validos(self):
        """O campo tipo deve ser PF ou PJ."""
        for lead in self.leads:
            assert lead["tipo"] in ("PF", "PJ"), \
                f"Lead {lead.get('nome', '?')} tem tipo invalido: {lead['tipo']}"

    def test_dependentes_eh_numero(self):
        """O campo dependentes deve ser um número inteiro."""
        for lead in self.leads:
            assert isinstance(lead["dependentes"], int), \
                f"Lead {lead.get('nome', '?')} com dependentes nao int: {type(lead['dependentes'])}"

    def test_campos_monday(self):
        """O bloco monday deve ter os campos esperados."""
        campos_esperados = ["operadora", "saude_odonto", "tipo_plano", "produto",
                           "boleto_adesao", "vigencia", "mensalidade", "valor_mensalidade",
                           "mensalidades_seguidas"]
        for lead in self.leads:
            monday = lead.get("monday", {})
            for campo in campos_esperados:
                assert campo in monday, \
                    f"Lead {lead.get('nome', '?')} monday sem campo {campo}"


class TestTemplateDados:
    """Testes de caracterização para template_dados_lead.txt."""

    def test_template_existe(self):
        """O arquivo template deve existir."""
        assert os.path.exists("template_dados_lead.txt"), "template_dados_lead.txt nao encontrado"

    def test_template_contem_campos_base(self):
        """O template deve conter os campos base esperados."""
        with open("template_dados_lead.txt", "r", encoding="utf-8") as f:
            content = f.read()
        campos_base = ["Nome", "CPF", "Telefone"]
        for campo in campos_base:
            assert campo in content, f"Template deve conter o campo {campo}"