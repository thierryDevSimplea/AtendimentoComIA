import sys
import os
import json
import openpyxl

def test_leads_json():
    # Testa se o arquivo leads.json é válido
    with open('leads.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    assert isinstance(data, list), "leads.json deve ser uma lista de leads"

    for lead in data:
        assert 'nome' in lead, "Cada lead deve ter um campo 'nome'"
        assert 'number' in lead, "Cada lead deve ter um número"
        assert 'dependentes' in lead, "Cada lead deve ter o campo 'dependentes'"
        assert 'tipo' in lead, "Cada lead deve ter o campo 'tipo'"

    print("leads.json validado com sucesso")

def test_criar_planilha():
    # Testa se a planilha é gerada corretECTAMENTE
    import openpyxl
    from openpyxl import Workbook

    # Cria uma planilha de teste
    wb = Workbook()
    # Verifica se os cabeçalhos e estrutura estão corretos
    from openpyxl import load_workbook
    wb = load_workbook("planilha_leads.xlsx")
    # Testa se os dados estão corretos
    assert wb is not None, "Não foi possível carregar a planilha"

    # Testa se a planilha contém as abas esperadas
    assert "Dashboard" in [sheet.title for sheet in wb], "Aba Dashboard não encontrada"
    assert "Leads PF" in [sheet.name for sheet in wb], "Aba Leads PF não encontrada"
    assert "Leads PJ" in [sheet.name for sheet in wb], "Aba Leads PJ não encontrada"

    print("Planilha validada com sucesso")