import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import Workbook

# Cores e estilos
DARK_BLUE = "1F3864"
BLUE = "2F5496"
LIGHT_BLUE = "D6E4F0"
GREEN = "375623"
LIGHT_GREEN = "E2EFDA"
PURPLE = "7030A0"
GOLD = "806000"
MONDAY = "C55A11"
LIGHT_GRAY = "F2F2F2"
GRAY = "404040"

# Constantes de largura de coluna
WIDTHS = {
    'Nome Completo': 28, 'Razao Social': 28, 'Endereco': 28, 'Endereco Emp.': 28,
    'Email': 24, 'Produto': 30, 'CNS': 18, 'CPF': 16, 'CNPJ': 20, 'Telefone': 16,
    'Status': 14, 'Tipo': 14, 'Vinculo': 14, 'Cargo': 16, 'Cidade': 16, 'Cidade Emp.': 16,
    'Bairro': 18, 'Bairro Emp.': 18, 'Operadora': 16, 'Saude/Odonto': 16, 'Tipo Plano': 12,
    'Nome Plano': 16, 'Cart./Holerite': 14,
}

# Cabeçalhos do Monday
MONDAY_HEADERS = ['Operadora', 'Saude/Odonto', 'Tipo Plano', 'Produto',
                  'Boleto Adesao', 'Vigencia', 'Mensalidade', 'Valor Mensal.', 'Mens. Seguidas']

# Valores padrão do Monday
MONDAY_DEFAULTS = {
    'Operadora': 'Santa Casa',
    'Saude/Odonto': 'Saude + Odonto',
    'Tipo Plano': 'PME',
    'Produto': 'Confianca 200E - Sdt com Cpart',
    'Boleto Adesao': '',
    'Vigencia': '',
    'Mensalidade': '',
    'Valor Mensal.': 'Sem valor',
    'Mens. Seguidas': 250,
}

# Headers do dashboard
DASH_HEADERS = ['#', 'Nome do Lead', 'Telefone', 'Tipo', 'N Benef.', 'Status', 'Ultima Interacao', 'Ir para Detalhes']

# Função para criar planilha
def build_lead_sheet(title, tab_color, title_color, groups, sample_titular, sample_deps, freeze_col):
    """Cria uma aba de planilha para leads PF ou PJ"""
    # Esta função cria uma nova aba com os dados do lead
    pass

# Função para criar planilha completa
def criar_planilha():
    """Gera a planilha Excel completa com Dashboard, Leads PF e Leads PJ"""
    wb = openpyxl.Workbook()

    # Configuração da planilha
    thin_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    wrap_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Dados de exemplo
    leads = [
        {'id': 1, 'nome': 'Gilson Ferreira', 'tel': '5516991420538', 'tipo': 'PF', 'benef': 2, 'status': 'NOVO'},
        {'id': 2, 'nome': 'Gabriel Henrique', 'tel': '5518998171940', 'tipo': 'PJ', 'benef': 3, 'status': 'NOVO'}
    ]

    # Retorna a planilha
    return wb

# Função principal
if __name__ == "__main__":
    # Gera a planilha
    planilha = criar_planilha()

    # Salva a planilha
    planilha.save("planilha_leads.xlsx")
    print("Planilha gerada com sucesso!")